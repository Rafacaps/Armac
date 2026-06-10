from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
import datetime
import io
import json

app = Flask(__name__)
CORS(app)

# ── Helpers ─────────────────────────────────────────────────
def fill(hex_): return PatternFill('solid', fgColor=hex_)
def font(bold=False, size=11, color='000000', italic=False):
    return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)
def border_thin():
    s = Side(style='thin', color='D1D5DB')
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, size=11, color='000000',
             bg=None, h_align='left', v_align='center', wrap=False,
             italic=False, border=False, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name='Arial', bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    if bg: c.fill = PatternFill('solid', fgColor=bg)
    if border: c.border = border_thin()
    if num_format: c.number_format = num_format
    return c

def norm_tipo(t):
    if not t: return 'Outros'
    u = t.strip().lower()
    if 'eletric' in u: return 'Elétrica'
    if 'hidraul' in u or 'esgoto' in u: return 'Hidráulica'
    if 'marcen' in u: return 'Marcenaria'
    if 'civil' in u: return 'Civil'
    if 'estrut' in u: return 'Estrutura'
    if 'detet' in u or 'dedet' in u: return 'Dedetização'
    return t.strip()

def norm_status(s):
    if not s: return 'Em Aberto'
    u = s.strip().upper()
    if u in ('CONCLUIDO','CONCLUÍDO') or u.startswith('RESOLVIDO') or u.startswith('EU MESMO') or u.startswith('O PROPRIO'):
        return 'Concluído'
    return 'Em Aberto'

# ── Cores ────────────────────────────────────────────────────
NAVY='0A1628'; BLUE='1E3A8A'; BLUE2='2563EB'
GREEN='059669'; GREEN2='D1FAE5'; RED='DC2626'; RED2='FEE2E2'
YELLOW='D97706'; YELLOW2='FEF3C7'; PURPLE='6366F1'; PURPLE2='EEF2FF'
GRAY1='F8FAFC'; GRAY2='E2E8F0'; GRAY3='94A3B8'; WHITE='FFFFFF'

def gerar_excel(aloj, oss):
    wb = Workbook()
    wb.remove(wb.active)

    ocupados = [a for a in aloj if a.get('nome','').strip() and a.get('status') != 'EM MANUTENÇÃO']
    vagos    = [a for a in aloj if not a.get('nome','').strip() and a.get('status') != 'EM MANUTENÇÃO']
    manut    = [a for a in aloj if a.get('status') == 'EM MANUTENÇÃO']
    homens   = [a for a in ocupados if 'masc' in (a.get('genero') or '').lower()]
    mulheres = [a for a in ocupados if 'fem'  in (a.get('genero') or '').lower()]
    total    = len(aloj)
    taxa     = round(len(ocupados)/total*100) if total else 0
    locs     = sorted(set(a['localizacao'] for a in aloj if a.get('localizacao')))
    letras   = sorted(set(a['letra'] for a in aloj if a.get('letra')))

    # ══ ABA 1 — PAINEL EXECUTIVO ══════════════════════════════
    ws1 = wb.create_sheet('📊 Painel Executivo')
    ws1.sheet_view.showGridLines = False
    for col_l, w in zip(['A','B','C','D','E','F','G','H','I'],
                        [2, 22, 16, 16, 16, 16, 16, 16, 2]):
        ws1.column_dimensions[col_l].width = w

    # Cabeçalho
    for r in range(1,5):
        for c in range(1,10):
            ws1.cell(row=r, column=c).fill = fill(NAVY)
    ws1.merge_cells('B1:H1')
    set_cell(ws1,1,2,'ARMAC | RELATÓRIO GERENCIAL DE ALOJAMENTO',bold=True,size=16,color=WHITE,bg=NAVY,h_align='center')
    ws1.row_dimensions[1].height = 32
    ws1.merge_cells('B2:H2')
    set_cell(ws1,2,2,'MINERAÇÃO TABOCA',size=11,color='94A3B8',bg=NAVY,h_align='center')
    ws1.merge_cells('B3:D3')
    set_cell(ws1,3,2,f'Gerado em: {datetime.datetime.now().strftime("%d/%m/%Y às %H:%M")}',italic=True,size=10,color='CBD5E1',bg=NAVY)
    ws1.merge_cells('E3:H3')
    set_cell(ws1,3,5,f'Ref.: {datetime.datetime.now().strftime("%B/%Y").upper()}',italic=True,size=10,color='CBD5E1',bg=NAVY,h_align='right')
    ws1.row_dimensions[3].height = 20
    ws1.row_dimensions[4].height = 8

    # KPIs
    kpis = [
        ('B','TOTAL DE LEITOS',total,'capacidade total',BLUE2,PURPLE2),
        ('C','OCUPADOS',len(ocupados),f'{taxa}% da cap.',GREEN,GREEN2),
        ('D','VAGOS',len(vagos),'disponíveis',PURPLE,PURPLE2),
        ('E','MANUTENÇÃO',len(manut),'indisponíveis',RED,RED2),
        ('F','HOMENS',len(homens),f'{round(len(homens)/len(ocupados)*100) if ocupados else 0}% aloj.',BLUE2,'DBEAFE'),
        ('G','MULHERES',len(mulheres),f'{round(len(mulheres)/len(ocupados)*100) if ocupados else 0}% aloj.','EC4899','FCE7F3'),
        ('H','CAP. MÁXIMA',total*2,'2 escalas/leito',YELLOW,YELLOW2),
    ]
    ws1.row_dimensions[5].height = 8
    for col_l,titulo,valor,sub,cor,bg_cor in kpis:
        col = ord(col_l)-64
        for r in range(6,12):
            ws1.cell(row=r,column=col).fill = fill(bg_cor)
            ws1.cell(row=r,column=col).border = border_thin()
        ws1.merge_cells(f'{col_l}6:{col_l}6')
        set_cell(ws1,6,col,titulo,bold=True,size=8,color=cor,h_align='center',bg=bg_cor)
        ws1.merge_cells(f'{col_l}7:{col_l}9')
        set_cell(ws1,7,col,valor,bold=True,size=22,color=cor,h_align='center',v_align='center',bg=bg_cor)
        set_cell(ws1,10,col,sub,size=8,color=GRAY3,h_align='center',bg=bg_cor,italic=True)
        set_cell(ws1,11,col,'',bg=bg_cor)
    for r,h in [(6,18),(7,28),(8,10),(9,10),(10,16),(11,6)]:
        ws1.row_dimensions[r].height = h

    # Tabela localizações
    ws1.row_dimensions[12].height = 14
    ws1.merge_cells('B13:I13')
    set_cell(ws1,13,2,'  📍  RESUMO POR LOCALIZAÇÃO',bold=True,size=12,color=WHITE,bg=NAVY)
    ws1.row_dimensions[13].height = 24
    ws1.column_dimensions['I'].width = 12
    for ci,h in enumerate(['Localização','Leitos','Ocupados','Vagos','Manutenção','Homens','Mulheres','Ocup.%']):
        set_cell(ws1,14,2+ci,h,bold=True,size=10,color=WHITE,bg=BLUE,
                 h_align='center' if ci>0 else 'left',border=True)
    ws1.row_dimensions[14].height = 20

    for li,loc in enumerate(locs):
        sub  = [a for a in aloj if a.get('localizacao')==loc]
        ocup = [a for a in sub if a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
        vag  = [a for a in sub if not a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
        man  = [a for a in sub if a.get('status')=='EM MANUTENÇÃO']
        hom  = [a for a in ocup if 'masc' in (a.get('genero') or '').lower()]
        mul  = [a for a in ocup if 'fem'  in (a.get('genero') or '').lower()]
        pct  = round(len(ocup)/len(sub)*100) if sub else 0
        bg_r = GRAY1 if li%2==0 else WHITE
        vals = [loc,len(sub),len(ocup),len(vag),len(man),len(hom),len(mul),f'{pct}%']
        for ci,v in enumerate(vals):
            cor_v = ('000000' if ci<2 else GREEN if ci==2 else PURPLE if ci==3 else
                     RED if ci==4 else BLUE2 if ci==5 else 'EC4899' if ci==6 else
                     (RED if pct>=90 else YELLOW if pct>=70 else GREEN))
            set_cell(ws1,15+li,2+ci,v,bold=(ci==0),size=10,color=cor_v,
                     bg=bg_r,h_align='center' if ci>0 else 'left',border=True)
        ws1.row_dimensions[15+li].height = 18

    tr = 15+len(locs)
    for ci,v in enumerate(['TOTAL GERAL',total,len(ocupados),len(vagos),len(manut),len(homens),len(mulheres),f'{taxa}%']):
        set_cell(ws1,tr,2+ci,v,bold=True,size=10,color=WHITE,bg=NAVY,
                 h_align='center' if ci>0 else 'left',border=True)
    ws1.row_dimensions[tr].height = 22
    

    # ══ ABA 2 — POR ESCALA ════════════════════════════════════
    ws2 = wb.create_sheet('🔤 Por Escala')
    ws2.sheet_view.showGridLines = False
    for col_l,w in zip(['A','B','C','D','E','F','G','H'],[2,14,16,20,14,14,14,2]):
        ws2.column_dimensions[col_l].width = w
    ws2.merge_cells('B1:G1')
    set_cell(ws2,1,2,'ALOJADOS POR ESCALA (LETRA)',bold=True,size=14,color=WHITE,bg=NAVY,h_align='center')
    ws2.row_dimensions[1].height = 30
    ws2.row_dimensions[2].height = 10

    for ci,h in enumerate(['Escala','Cap. Leitos','Cap. Pessoas (x2)','Alojados','Vagas','Ocup. %']):
        set_cell(ws2,3,2+ci,h,bold=True,size=10,color=WHITE,bg=BLUE,h_align='center',border=True)
    ws2.row_dimensions[3].height = 22

    for li,l in enumerate(letras):
        sub  = [a for a in aloj if a.get('letra')==l]
        ocup = [a for a in sub if a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
        cap  = len(sub)
        pct  = round(len(ocup)/cap*100) if cap else 0
        cor_pct = RED if pct>=90 else (YELLOW if pct>=70 else GREEN)
        bg_r = GRAY1 if li%2==0 else WHITE
        for ci,v in enumerate([l,cap,cap*2,len(ocup),cap-len(ocup),f'{pct}%']):
            set_cell(ws2,4+li,2+ci,v,bold=(ci==0 or ci==5),size=11,
                     color=cor_pct if ci==5 else '000000',
                     bg=bg_r,h_align='center',border=True)
        ws2.row_dimensions[4+li].height = 22

    # Gráfico barras
    chart1 = BarChart()
    chart1.type = 'col'; chart1.grouping = 'clustered'
    chart1.title = 'Capacidade vs Alojados por Escala'
    chart1.style = 10; chart1.width = 20; chart1.height = 13
    cats = Reference(ws2,min_col=2,min_row=4,max_row=4+len(letras)-1)
    d_cap  = Reference(ws2,min_col=3,min_row=3,max_row=4+len(letras)-1)
    d_aloj = Reference(ws2,min_col=5,min_row=3,max_row=4+len(letras)-1)
    chart1.add_data(d_cap,titles_from_data=True)
    chart1.add_data(d_aloj,titles_from_data=True)
    chart1.set_categories(cats)
    chart1.series[0].graphicalProperties.solidFill = '3B82F6'
    chart1.series[1].graphicalProperties.solidFill = '10B981'
    ws2.add_chart(chart1,'B'+str(6+len(letras)))

    # Gráfico pizza
    chart2 = PieChart()
    chart2.title = 'Distribuição por Escala'; chart2.style = 10
    chart2.width = 14; chart2.height = 13
    d_pie = Reference(ws2,min_col=5,min_row=3,max_row=4+len(letras)-1)
    chart2.add_data(d_pie,titles_from_data=True)
    chart2.set_categories(cats)
    ws2.add_chart(chart2,'G'+str(6+len(letras)))

    # ══ ABA 3 — OS E MANUTENÇÃO ═══════════════════════════════
    ws3 = wb.create_sheet('🔧 OS e Manutenção')
    ws3.sheet_view.showGridLines = False
    for col_l,w in zip(['A','B','C','D','E','F','G','H','I'],[2,8,14,25,16,45,16,14,2]):
        ws3.column_dimensions[col_l].width = w
    ws3.merge_cells('B1:H1')
    set_cell(ws3,1,2,'OS E CHAMADOS DE MANUTENÇÃO',bold=True,size=14,color=WHITE,bg=NAVY,h_align='center')
    ws3.row_dimensions[1].height = 30
    ws3.row_dimensions[2].height = 10

    # Todas as OS (sem filtrar por local — o campo pode estar vazio)
    os_ab = [o for o in oss if norm_status(o.get('status'))=='Em Aberto']
    os_co = [o for o in oss if norm_status(o.get('status'))=='Concluído']
    kpis3 = [('B','TOTAL OS',len(oss),NAVY,'F1F5F9'),
             ('D','EM ABERTO',len(os_ab),YELLOW,YELLOW2),
             ('F','CONCLUÍDAS',len(os_co),GREEN,GREEN2),
             ('H','RESOLVIDAS',f'{round(len(os_co)/len(oss)*100) if oss else 0}%',BLUE2,'DBEAFE')]
    for col_l,titulo,valor,cor,bg_cor in kpis3:
        col=ord(col_l)-64
        for r in range(3,8):
            for dc in range(2):
                ws3.cell(row=r,column=col+dc).fill=fill(bg_cor)
                ws3.cell(row=r,column=col+dc).border=border_thin()
        ws3.merge_cells(f'{col_l}3:{get_column_letter(col+1)}3')
        set_cell(ws3,3,col,titulo,bold=True,size=9,color=cor,h_align='center',bg=bg_cor)
        ws3.merge_cells(f'{col_l}4:{get_column_letter(col+1)}6')
        set_cell(ws3,4,col,valor,bold=True,size=22,color=cor,h_align='center',v_align='center',bg=bg_cor)
        ws3.merge_cells(f'{col_l}7:{get_column_letter(col+1)}7')
        set_cell(ws3,7,col,'',bg=bg_cor)
    for r in [3,4,5,6,7,8]: ws3.row_dimensions[r].height = 20

    for ci,h in enumerate(['#','Data','Local','Tipo','Descrição','Status','Conclusão']):
        set_cell(ws3,9,2+ci,h,bold=True,size=10,color=WHITE,bg=BLUE,h_align='center',border=True)
    ws3.row_dimensions[9].height = 22

    for oi,o in enumerate(sorted(oss,key=lambda x:x.get('data_abertura','') or '')):
        st  = norm_status(o.get('status'))
        bg_r = YELLOW2 if st=='Em Aberto' else (GREEN2 if oi%2==0 else WHITE)
        try: dt_fmt  = datetime.datetime.strptime((o.get('data_abertura') or '')[:10],'%Y-%m-%d').strftime('%d/%m/%Y')
        except: dt_fmt = '—'
        try: dtc_fmt = datetime.datetime.strptime((o.get('data_conclusao') or '')[:10],'%Y-%m-%d').strftime('%d/%m/%Y')
        except: dtc_fmt = '—'
        vals = [oi+1,dt_fmt,(o.get('local') or '')[:30],norm_tipo(o.get('tipo') or ''),
                (o.get('descricao') or '')[:60],st,dtc_fmt]
        for ci,v in enumerate(vals):
            cor_v = RED if (st=='Em Aberto' and ci==5) else (GREEN if ci==5 else '000000')
            set_cell(ws3,10+oi,2+ci,v,size=10,color=cor_v,bg=bg_r,
                     h_align='left' if ci==4 else 'center',border=True,wrap=(ci==4))
        ws3.row_dimensions[10+oi].height = 20

    # Dados para gráficos OS
    base_r = max(12+len(oss), 15)
    ws3.cell(row=base_r,column=2).value='Status'; ws3.cell(row=base_r,column=3).value='Qtd'
    ws3.cell(row=base_r+1,column=2).value='Em Aberto'; ws3.cell(row=base_r+1,column=3).value=len(os_ab)
    ws3.cell(row=base_r+2,column=2).value='Concluídas'; ws3.cell(row=base_r+2,column=3).value=len(os_co)
    c_pie=PieChart(); c_pie.title='Status das OS'; c_pie.style=10; c_pie.width=14; c_pie.height=10
    c_pie.add_data(Reference(ws3,min_col=3,min_row=base_r,max_row=base_r+2),titles_from_data=True)
    c_pie.set_categories(Reference(ws3,min_col=2,min_row=base_r+1,max_row=base_r+2))
    ws3.add_chart(c_pie,'E'+str(base_r))

    tipos_map={}
    for o in oss: t=norm_tipo(o.get('tipo','')); tipos_map[t]=tipos_map.get(t,0)+1
    t_row=base_r
    ws3.cell(row=t_row,column=6).value='Tipo'; ws3.cell(row=t_row,column=7).value='Qtd'
    for ti,(t,v) in enumerate(sorted(tipos_map.items(),key=lambda x:-x[1])):
        ws3.cell(row=t_row+1+ti,column=6).value=t; ws3.cell(row=t_row+1+ti,column=7).value=v
    c_bar=BarChart(); c_bar.type='bar'; c_bar.title='OS por Tipo'; c_bar.style=10
    c_bar.width=18; c_bar.height=10
    c_bar.add_data(Reference(ws3,min_col=7,min_row=t_row,max_row=t_row+len(tipos_map)),titles_from_data=True)
    c_bar.set_categories(Reference(ws3,min_col=6,min_row=t_row+1,max_row=t_row+len(tipos_map)))
    c_bar.series[0].graphicalProperties.solidFill='6366F1'
    ws3.add_chart(c_bar,'I'+str(t_row))

    # ══ ABA 4 — LISTA COMPLETA ════════════════════════════════
    ws4 = wb.create_sheet('📋 Lista Completa')
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'B3'
    for col_l,w in zip(['A','B','C','D','E','F','G','H','I','J','K'],
                       [2,8,8,35,16,32,22,10,10,12,14]):
        ws4.column_dimensions[col_l].width = w
    ws4.merge_cells('B1:K1')
    set_cell(ws4,1,2,'LISTA COMPLETA DE ALOJAMENTOS — ARMAC TABOCA',bold=True,size=13,color=WHITE,bg=NAVY,h_align='center')
    ws4.row_dimensions[1].height = 28
    for ci,h in enumerate(['Item','Letra','Localização','Quarto','Nome','Função','Gênero','Turno','Situação','Status']):
        set_cell(ws4,2,2+ci,h,bold=True,size=10,color=WHITE,bg=BLUE,h_align='center',border=True)
    ws4.row_dimensions[2].height = 22

    for ai,a in enumerate(sorted(aloj,key=lambda x:(x.get('localizacao',''),x.get('quarto',''),x.get('item') or 0))):
        is_m=a.get('status')=='EM MANUTENÇÃO'; is_v=not a.get('nome','').strip()
        bg_r=RED2 if is_m else (PURPLE2 if is_v else (GRAY1 if ai%2==0 else WHITE))
        vals=[a.get('item',''),a.get('letra',''),a.get('localizacao',''),a.get('quarto',''),
              a.get('nome','') or '(vago)',a.get('funcao','') or '—',
              a.get('genero','') or '—',a.get('turno','') or '—',
              a.get('situacao','') or 'Normal',a.get('status','')]
        for ci,v in enumerate(vals):
            cor_v = RED if (is_m and ci==9) else (PURPLE if (is_v and ci==9) else (GREEN if ci==9 else '000000'))
            set_cell(ws4,3+ai,2+ci,v,size=9,color=cor_v,bg=bg_r,
                     h_align='left' if ci in [2,4,5] else 'center',border=True)
        ws4.row_dimensions[3+ai].height = 16

    # ══ ABA 5 — EM MANUTENÇÃO ════════════════════════════════
    ws5 = wb.create_sheet('🔴 Em Manutenção')
    ws5.sheet_view.showGridLines = False
    for col_l,w in zip(['A','B','C','D','E','F','G'],[2,8,8,35,20,35,35]):
        ws5.column_dimensions[col_l].width = w
    ws5.merge_cells('B1:G1')
    set_cell(ws5,1,2,f'LEITOS EM MANUTENÇÃO — {len(manut)} REGISTROS',bold=True,size=13,color=WHITE,bg=RED,h_align='center')
    ws5.row_dimensions[1].height = 28
    if not manut:
        ws5.merge_cells('B3:G3')
        set_cell(ws5,3,2,'✅ Nenhum leito em manutenção!',bold=True,size=12,color=GREEN,h_align='center')
    else:
        for ci,h in enumerate(['Item','Letra','Localização','Quarto','Colaborador','Observação']):
            set_cell(ws5,2,2+ci,h,bold=True,size=10,color=WHITE,bg=RED,h_align='center',border=True)
        ws5.row_dimensions[2].height = 22
        for mi,a in enumerate(manut):
            for ci,v in enumerate([a.get('item',''),a.get('letra',''),a.get('localizacao',''),
                                    a.get('quarto',''),a.get('nome','') or '(sem colaborador)',a.get('obs','') or '—']):
                set_cell(ws5,3+mi,2+ci,v,size=10,bg=RED2,
                         h_align='left' if ci in [2,4,5] else 'center',border=True)
            ws5.row_dimensions[3+mi].height = 18

    # Garantir !ref válido em todas as abas
    for ws in wb.worksheets:
        if ws.max_row and ws.max_column:
            ws.sheet_view.showGridLines = False
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── Rotas ────────────────────────────────────────────────────
@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok', 'service': 'ARMAC Relatorio Server'})

@app.route('/gerar-excel', methods=['POST'])
def gerar():
    try:
        body = request.get_json()
        aloj = body.get('alojamento', [])
        oss  = body.get('os_chamados', [])
        if not aloj:
            return jsonify({'erro': 'Dados de alojamento ausentes'}), 400
        buf = gerar_excel(aloj, oss)
        ts  = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'ARMAC_Relatorio_Alojamento_{ts}.xlsx'
        )
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)


def gerar_excel_pbi(aloj, oss, efs=None, mobs=None):
    """Gera Excel estruturado para Power BI"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import datetime, io

    wb = Workbook()
    wb.remove(wb.active)

    NAVY='0F2644'; BLUE='1A56DB'; WHITE='FFFFFF'; GRAY1='F3F4F6'; GRAY2='E5E7EB'

    def hdr(ws, row, cols):
        for ci, h in enumerate(cols):
            c = ws.cell(row=row, column=ci+1, value=h)
            c.font = Font(name='Calibri', bold=True, size=10, color=WHITE)
            c.fill = PatternFill('solid', fgColor=NAVY)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(bottom=Side(style='medium', color=BLUE))
        ws.row_dimensions[row].height = 20

    def row_d(ws, row, vals, alt=False):
        bg = GRAY1 if alt else WHITE
        for ci, v in enumerate(vals):
            c = ws.cell(row=row, column=ci+1, value=v)
            c.font = Font(name='Calibri', size=9)
            c.fill = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = Border(bottom=Side(style='thin', color=GRAY2), right=Side(style='thin', color=GRAY2))
        ws.row_dimensions[row].height = 15

    def bloco_tipo(loc):
        bl = (loc or '').lower()
        if 'container' in bl: return 'Container'
        elif 'vila' in bl or 'repub' in bl: return 'Vila/República'
        elif 'bloco' in bl and 'metal' in bl: return 'Bloco Metálico'
        elif 'bloco' in bl and 'concret' in bl: return 'Bloco Concreto'
        elif 'uhe' in bl or 'usina' in bl: return 'UHE'
        elif 'vias' in bl: return 'Bloco Vias'
        return 'Outros'

    def ng(g):
        if not g: return 'Não Informado'
        u = g.lower()
        if 'masc' in u or u=='m': return 'Masculino'
        if 'fem' in u or u=='f': return 'Feminino'
        return 'Não Informado'

    def ns(a):
        if a.get('status')=='EM MANUTENÇÃO': return 'Em Manutenção'
        if a.get('nome','').strip(): return 'Ocupado'
        return 'Vago'

    # fAlojamento
    ws1 = wb.create_sheet('fAlojamento')
    ws1.sheet_view.showGridLines = False
    ws1.freeze_panes = 'A2'
    cols1=['ID','Item','Letra_Escala','Localizacao','Bloco_Tipo','Quarto','Nome_Colaborador',
           'Funcao','Genero','Turno','Situacao','Status_Leito','Status_Categoria','Matricula','Area']
    hdr(ws1, 1, cols1)
    for ci,w in enumerate([6,6,12,35,18,18,35,25,14,10,12,16,16,12,15]):
        ws1.column_dimensions[get_column_letter(ci+1)].width = w
    for ai,a in enumerate(sorted(aloj, key=lambda x:x.get('item') or 0)):
        row_d(ws1, ai+2, [a.get('id',''),a.get('item',''),a.get('letra','') or '',
            a.get('localizacao','') or '',bloco_tipo(a.get('localizacao','')),
            a.get('quarto','') or '',a.get('nome','') or '',a.get('funcao','') or '',
            ng(a.get('genero','')),a.get('turno','') or '',a.get('situacao','') or 'Normal',
            a.get('status','') or '',ns(a),a.get('matricula','') or '',a.get('area','') or ''], alt=ai%2==1)
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(cols1))}1'

    # fOS_Chamados
    ws2 = wb.create_sheet('fOS_Chamados')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A2'
    cols2=['ID','Data_Abertura','Tipo_Servico','Tipo_Normalizado','Local','Descricao',
           'Status_Original','Status_Normalizado','Data_Conclusao','Dias_Resolucao','Responsavel']
    hdr(ws2, 1, cols2)
    for ci,w in enumerate([6,14,18,16,30,45,20,16,14,14,20]):
        ws2.column_dimensions[get_column_letter(ci+1)].width = w
    for oi,o in enumerate(oss):
        st = norm_status(o.get('status'))
        try:
            dt_ab = datetime.datetime.strptime((o.get('data_abertura','') or '')[:10],'%Y-%m-%d')
            if st=='Concluído' and o.get('data_conclusao'):
                dt_co = datetime.datetime.strptime((o.get('data_conclusao','') or '')[:10],'%Y-%m-%d')
                dias = (dt_co-dt_ab).days
            else:
                dias = (datetime.datetime.now()-dt_ab).days
        except: dias = None
        row_d(ws2, oi+2, [o.get('id',''), (o.get('data_abertura','') or '')[:10],
            o.get('tipo','') or '', norm_tipo(o.get('tipo','')), o.get('local','') or '',
            o.get('descricao','') or '', o.get('status','') or '', st,
            (o.get('data_conclusao','') or '')[:10], dias, o.get('ass_adm','') or ''], alt=oi%2==1)
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(cols2))}1'

    # dLocalizacao
    ws3 = wb.create_sheet('dLocalizacao')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A2'
    cols3=['Localizacao','Bloco_Tipo','Total_Leitos','Ocupados','Vagos','Em_Manutencao',
           'Homens','Mulheres','Nao_Informado','Taxa_Ocupacao_Pct','Capacidade_Max']
    hdr(ws3, 1, cols3)
    locs = sorted(set(a['localizacao'] for a in aloj if a.get('localizacao')))
    for li,loc in enumerate(locs):
        sub  = [a for a in aloj if a.get('localizacao')==loc]
        ocup = [a for a in sub if a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
        vag  = [a for a in sub if not a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
        man  = [a for a in sub if a.get('status')=='EM MANUTENÇÃO']
        hom  = [a for a in ocup if 'masc' in (a.get('genero') or '').lower()]
        mul  = [a for a in ocup if 'fem'  in (a.get('genero') or '').lower()]
        ni   = len(ocup)-len(hom)-len(mul)
        pct  = round(len(ocup)/len(sub)*100) if sub else 0
        row_d(ws3, li+2, [loc, bloco_tipo(loc), len(sub), len(ocup), len(vag), len(man),
                           len(hom), len(mul), ni, pct, len(sub)*2], alt=li%2==1)

    # dEscala
    ws4 = wb.create_sheet('dEscala')
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'A2'
    cols4=['Letra_Escala','Total_Leitos','Capacidade_Pessoas','Alojados','Vagas','Taxa_Ocupacao_Pct']
    hdr(ws4, 1, cols4)
    letras = sorted(set(a['letra'] for a in aloj if a.get('letra')))
    for li,l in enumerate(letras):
        sub  = [a for a in aloj if a.get('letra')==l]
        ocup = [a for a in sub if a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
        pct  = round(len(ocup)/len(sub)*100) if sub else 0
        row_d(ws4, li+2, [l, len(sub), len(sub)*2, len(ocup), len(sub)-len(ocup), pct], alt=li%2==1)

    # dResumo_KPIs
    ws5 = wb.create_sheet('dResumo_KPIs')
    ws5.sheet_view.showGridLines = False
    cols5=['Indicador','Valor','Unidade','Categoria']
    hdr(ws5, 1, cols5)
    ws5.column_dimensions['A'].width = 30
    ws5.column_dimensions['B'].width = 12
    ws5.column_dimensions['C'].width = 12
    ws5.column_dimensions['D'].width = 16
    total = len(aloj)
    ocupados = [a for a in aloj if a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
    vagos    = [a for a in aloj if not a.get('nome','').strip() and a.get('status')!='EM MANUTENÇÃO']
    manut    = [a for a in aloj if a.get('status')=='EM MANUTENÇÃO']
    homens   = [a for a in ocupados if 'masc' in (a.get('genero') or '').lower()]
    mulheres = [a for a in ocupados if 'fem'  in (a.get('genero') or '').lower()]
    taxa     = round(len(ocupados)/total*100) if total else 0
    os_ab    = [o for o in oss if norm_status(o.get('status'))=='Em Aberto']
    os_co    = [o for o in oss if norm_status(o.get('status'))=='Concluído']
    kpis = [
        ('Total de Leitos',total,'leitos','Alojamento'),
        ('Leitos Ocupados',len(ocupados),'leitos','Alojamento'),
        ('Leitos Vagos',len(vagos),'leitos','Alojamento'),
        ('Em Manutenção',len(manut),'leitos','Alojamento'),
        ('Taxa de Ocupação',taxa,'%','Alojamento'),
        ('Capacidade Máxima',total*2,'pessoas','Alojamento'),
        ('Masculino',len(homens),'pessoas','Gênero'),
        ('Feminino',len(mulheres),'pessoas','Gênero'),
        ('Total OS',len(oss),'OS','Manutenção'),
        ('OS Em Aberto',len(os_ab),'OS','Manutenção'),
        ('OS Concluídas',len(os_co),'OS','Manutenção'),
        ('Taxa Resolução OS',round(len(os_co)/len(oss)*100) if oss else 0,'%','Manutenção'),
    ]
    for ki,(ind,val,uni,cat) in enumerate(kpis):
        row_d(ws5, ki+2, [ind,val,uni,cat], alt=ki%2==1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route('/gerar-powerbi', methods=['POST'])
def gerar_pbi():
    try:
        body  = request.get_json()
        aloj  = body.get('alojamento', [])
        oss   = body.get('os_chamados', [])
        efs   = body.get('efetivo', [])
        mobs  = body.get('colaboradores', [])
        if not aloj:
            return jsonify({'erro': 'Dados de alojamento ausentes'}), 400
        buf = gerar_excel_pbi(aloj, oss, efs, mobs)
        ts  = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'ARMAC_PowerBI_{ts}.xlsx')
    except Exception as e:
        return jsonify({'erro': str(e)}), 500
